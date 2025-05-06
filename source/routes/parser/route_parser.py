import os
import csv
from flask import request, jsonify, Blueprint, send_file
from werkzeug.utils import secure_filename
from source.models.user_model import MongoDB
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from io import BytesIO
from datetime import datetime

from source.utils.utils import token_required

parser_bp = Blueprint("parser_bp",__name__)

UPLOAD_DIR = "uploaded_files"

# Define your header mapping here
HEADER_MAP = {
    "OldName1": "new_field1",
    "OldName2": "new_field2",
    "OldName3": "new_field3",
}

@parser_bp.route('/upload_csv', methods=['POST'])
@token_required
def upload_csv_route():
    if 'file' not in request.files:
        return jsonify({"message": "CSV file is missing in the request"}), 400

    file = request.files['file']
    if not file.filename.endswith('.csv'):
        return jsonify({"message": "Only CSV files are allowed"}), 400

    user_email = request.user_email
    user_folder = os.path.join(UPLOAD_DIR, user_email)
    os.makedirs(user_folder, exist_ok=True)

    filename = secure_filename(file.filename)
    file_path = os.path.join(user_folder, filename)
    file.save(file_path)

    with open(file_path, newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        transformed_data = []

        for row in reader:
            transformed_row = {}
            for old_key, value in row.items():
                new_key = HEADER_MAP.get(old_key, old_key)  
                transformed_row[new_key] = value
            transformed_data.append(transformed_row)

    mongo = MongoDB()
    mongo.csv_data.insert_many(transformed_data)

    return jsonify({"message": "File uploaded and data inserted", "records_inserted": len(transformed_data)}), 200

@parser_bp.route('/download-excel', methods=['GET'])
def download_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"

    # Headers in desired order
    headers = [
        "Village",
        "Survey Number",
        "Range Number",
        "Division Register Number",
        "Range Division Number",
        "Incident Date",  
        "Incident Type",
        "Compensation Amount"
    ]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    date_validation = DataValidation(
        type="date",
        formula1="DATE(1900,1,1)",
        allow_blank=True,
        showInputMessage=True,
        promptTitle="Date Entry",
        prompt="Enter a valid date (e.g., 15-05-2024)."
    )
    ws.add_data_validation(date_validation)

    date_range = "F2:F1000"
    date_validation.add(date_range)

    for row in range(2, 1001):
        ws.cell(row=row, column=6).number_format = "DD-MM-YYYY"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name="template.xlsx"
    )